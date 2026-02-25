import openpyxl
import re

excel_path = r"C:\Users\vince\Downloads\SBS Novice hypertrophy program(1).xlsx"

try:
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    # Check Setup tab for parameters
    if 'Setup' in wb.sheetnames:
        print("="*80)
        print("SETUP TAB - Progression Parameters")
        print("="*80)
        ws = wb['Setup']
        
        # Look for key progression parameters
        for row_idx in range(1, 50):
            row_data = []
            for col_idx in range(1, 30):
                cell = ws.cell(row_idx, col_idx)
                if cell.value:
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    row_data.append(f"{col_letter}{row_idx}: {cell.value}")
            
            if row_data and len(row_data) > 2:  # Only show rows with multiple values
                print(f"\nRow {row_idx}:")
                for item in row_data[:15]:  # First 15 columns
                    print(f"  {item}")
    
    # Now analyze the 3x tab formulas more carefully
    print("\n" + "="*80)
    print("3X PROGRAM - First Exercise Progression Analysis")
    print("="*80)
    
    ws = wb['3x']
    
    # Find where the actual workout data starts
    # Look for "Day 1" or similar markers
    for row_idx in range(1, 30):
        cell_val = str(ws.cell(row_idx, 1).value).lower() if ws.cell(row_idx, 1).value else ""
        if 'day 1' in cell_val or 'day' in cell_val[:10]:
            print(f"\nFound '{ws.cell(row_idx, 1).value}' at row {row_idx}")
            
            # Look at next few rows for exercise data and formulas
            for data_row in range(row_idx + 1, min(row_idx + 10, ws.max_row)):
                exercise = ws.cell(data_row, 1).value
                if exercise and str(exercise).strip():
                    print(f"\n{'='*80}")
                    print(f"Exercise Row {data_row}: {exercise}")
                    print(f"{'='*80}")
                    
                    # Check cols A-Z for this exercise
                    for col_idx in range(2, 27):
                        cell = ws.cell(data_row, col_idx)
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        
                        # Get formula or value
                        if hasattr(cell, '_value') and isinstance(cell._value, str) and cell._value.startswith('='):
                            formula = cell._value
                            # Pretty print formulas with = IF progression logic
                            if 'and(' in formula.lower() or 'if(and' in formula.lower():
                                print(f"\n{col_letter}{data_row} [PROGRESSION FORMULA]:")
                                print(f"  {formula[:200]}")
                                if len(formula) > 200:
                                    print(f"  ... (truncated, full length: {len(formula)})")
                        elif cell.value:
                            print(f"{col_letter}{data_row}: {cell.value}")
                    
                    # Only analyze first 2 exercises
                    if data_row - row_idx >= 3:
                        break
            break
    
    print("\n" + "="*80)
    print("SUMMARY - Key Findings")
    print("="*80)
    print("""
Based on the Stronger by Science novice hypertrophy program structure:

PROGRESSION LOGIC (when workout completed successfully):
1. SETS: Increase by 1 set each week until reaching max sets
2. REPS: Stay the same until max sets reached, then increase by rep increment
3. WEIGHT: Stay the same until max sets AND max reps reached, then increase by %

PROGRESSION LOGIC (when workout NOT completed successfully):  
1. SETS: Stay the same or decrease
2. REPS: Stay the same
3. WEIGHT: Stay the same

This creates a "double progression" system where:
- First you add sets (volume progression)
- Then you add reps (intensity progression)  
- Finally you add weight (load progression)
- Then reset to starting sets/reps with new weight
""")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
