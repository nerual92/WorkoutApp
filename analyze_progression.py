import openpyxl
import json
from pathlib import Path

# Path to the Excel file
excel_path = r"C:\Users\vince\Downloads\SBS Novice hypertrophy program(1).xlsx"

try:
    # Load the workbook
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    print("Available sheets:", wb.sheetnames)
    print("\n" + "="*80 + "\n")
    
    # Analyze each program variant (3x, 4x, 5x)
    for sheet_name in ['3x', '4x', '5x']:
        if sheet_name in wb.sheetnames:
            print(f"\n{'='*80}")
            print(f"ANALYZING {sheet_name.upper()} PROGRAM")
            print(f"{'='*80}\n")
            
            ws = wb[sheet_name]
            
            # Find the header row (look for "Exercise", "Sets", "Reps", "Weight")
            header_row = None
            for row_idx in range(1, 20):
                row_values = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[row_idx]]
                if 'exercise' in row_values or 'sets' in row_values:
                    header_row = row_idx
                    break
            
            if not header_row:
                print(f"Could not find header row in {sheet_name}")
                continue
            
            # Get column headers
            headers = []
            for col_idx, cell in enumerate(ws[header_row], start=1):
                val = str(cell.value).strip() if cell.value else ""
                headers.append((col_idx, val))
            
            print("Column Headers:")
            for col_idx, header in headers:
                print(f"  Col {col_idx}: {header}")
            print()
            
            # Find Sets, Reps, Weight columns
            sets_cols = [col for col, h in headers if 'set' in h.lower() and 'set' in h.lower()[:4]]
            reps_cols = [col for col, h in headers if 'rep' in h.lower()]
            weight_cols = [col for col, h in headers if 'weight' in h.lower() or '%' in h.lower()]
            
            print(f"Sets columns: {sets_cols}")
            print(f"Reps columns: {reps_cols}")
            print(f"Weight columns: {weight_cols}")
            print()
            
            # Sample the first 20 weeks of data
            print("Sample Data (first 20 weeks):")
            print("-" * 100)
            
            # Print week by week
            for row_idx in range(header_row + 1, min(header_row + 21, ws.max_row + 1)):
                week_cell = ws.cell(row_idx, 1)
                week_val = week_cell.value
                
                if week_val:
                    print(f"\nWeek/Row {row_idx - header_row}: {week_val}")
                    
                    # Get formulas/values for Sets, Reps, Weight columns
                    if sets_cols:
                        for col in sets_cols[:3]:  # First 3 sets columns
                            cell = ws.cell(row_idx, col)
                            col_name = headers[col-1][1]
                            if hasattr(cell, 'value') and cell.value is not None:
                                # Try to get formula if it exists
                                formula = getattr(cell, '_value', None)
                                if isinstance(formula, str) and formula.startswith('='):
                                    print(f"  {col_name}: FORMULA: {formula[:100]}")
                                else:
                                    print(f"  {col_name}: {cell.value}")
                    
                    if reps_cols:
                        for col in reps_cols[:3]:  # First 3 reps columns
                            cell = ws.cell(row_idx, col)
                            col_name = headers[col-1][1]
                            if hasattr(cell, 'value') and cell.value is not None:
                                formula = getattr(cell, '_value', None)
                                if isinstance(formula, str) and formula.startswith('='):
                                    print(f"  {col_name}: FORMULA: {formula[:100]}")
                                else:
                                    print(f"  {col_name}: {cell.value}")
                    
                    if weight_cols:
                        for col in weight_cols[:3]:  # First 3 weight columns
                            cell = ws.cell(row_idx, col)
                            col_name = headers[col-1][1]
                            if hasattr(cell, 'value') and cell.value is not None:
                                formula = getattr(cell, '_value', None)
                                if isinstance(formula, str) and formula.startswith('='):
                                    print(f"  {col_name}: FORMULA: {formula[:100]}")
                                else:
                                    print(f"  {col_name}: {cell.value}")
            
            print("\n" + "="*80)
        else:
            print(f"Sheet '{sheet_name}' not found")
    
    print("\n\nANALYSIS COMPLETE")
    print("="*80)

except FileNotFoundError:
    print(f"Error: Excel file not found at {excel_path}")
    print("Please verify the file path is correct.")
except Exception as e:
    print(f"Error analyzing Excel file: {type(e).__name__}: {e}")
    import traceback
    traceback.print_exc()
