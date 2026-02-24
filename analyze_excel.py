import openpyxl
import re

wb = openpyxl.load_workbook('C:/Users/vince/Downloads/SBS Novice hypertrophy program(1).xlsx', data_only=False)
qs = wb['Quick Setup']

# Build Quick Setup lookup
qs_lookup = {}
for row in range(5, 27):
    exercise = qs[f'C{row}'].value
    muscle_group = qs[f'D{row}'].value
    if exercise:
        qs_lookup[row] = {'exercise': exercise, 'muscle_group': muscle_group}

def print_program(sheet_name, title):
    print(f'\n=== {title} ===')
    ws = wb[sheet_name]
    current_day = None
    for row in range(4, 60):
        col_a = ws[f'A{row}'].value
        if col_a == 'Accessories':
            print('  ---------- Optional Accessories ----------')
            continue
        if col_a and 'Day' in str(col_a):
            current_day = col_a
            print(f'\n{col_a}')
            continue
        if col_a and isinstance(col_a, str) and '!C' in col_a:
            match = re.search(r'C(\d+)', col_a)
            if match:
                qs_row = int(match.group(1))
                if qs_row in qs_lookup:
                    info = qs_lookup[qs_row]
                    sets = ws[f'C{row}'].value
                    reps = ws[f'D{row}'].value
                    mg = info['muscle_group']
                    ex = info['exercise']
                    print(f'  {mg:<40} (default: {ex})')

print_program('3x', '3X PROGRAM (3-day)')
print_program('4x', '4X PROGRAM (4-day)')
print_program('5x', '5X PROGRAM (5-day)')
