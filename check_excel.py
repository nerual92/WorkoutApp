import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\vince\Downloads\SBS Novice hypertrophy program(1).xlsx')
ws = wb['Quick Setup']

print("Checking data validations in Quick Setup:")
for dv in ws.data_validations.dataValidation:
    print(f"\nValidation formula: {dv.formula1}")
    print(f"Type: {dv.type}")
    print(f"Applies to: {dv.sqref}")

print("\n\nAll pec-dominant exercises found:")
for i in range(1, ws.max_row + 1):
    muscle_group = ws.cell(i, 4).value
    if muscle_group and 'pec' in str(muscle_group).lower():
        print(f"Row {i}: {ws.cell(i, 3).value}")
